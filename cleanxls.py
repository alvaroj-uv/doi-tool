from openpyxl import load_workbook
import re
wb = load_workbook(filename = 'JCR.xlsx')
sh = wb['db']
print(sh['D1'].value)

with open('./t-todasunique.txt', "w", encoding="utf-8") as tt:

    for row in range(2,21432):
        valor = sh.cell(row,4).value
        if valor:
            arr=valor.split(';')
            bu=';'.join([x[-4:] for x in valor.split(';')])
            bu=bu.replace('(', '').replace(')', '')
            bu=';'.join([x[-1:] for x in bu.split(';')])
            bu=bu.replace('A;','').replace('A', '')
            arr2=bu.split(';')
            arr2.sort(reverse=True)
            print(arr2)
        arrg=[]
        for x in range(1,9):
            arrg.append(sh.cell(row,x).value)
        if valor:
            arrg.append('Q'+arr2[0]+'')
        else:
            arrg.append('N/A')
        linea = '|'.join(arrg)
        tt.write(linea + '\n')


