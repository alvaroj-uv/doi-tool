import sqlite3
from sqlite3 import Error
import urllib.request
from openpyxl import load_workbook
import re

def prepare_file():

    wb = load_workbook(filename = 'JCR.xlsx')
    sh = wb['db']
    print(sh['D1'].value)

    with open('./t-todasunique.txt', "w", encoding="utf-8") as tt:

        for row in range(2,21432):
            valor = sh.cell(row,4).value
            if valor:
                arr=valor.split(';')
                bu=';'.join([x[-4:] for x in valor.split(';')])
                bu=bu.replace('(', '').replace(')', '')
                bu=';'.join([x[-1:] for x in bu.split(';')])
                bu=bu.replace('A;','').replace('A', '')
                arr2=bu.split(';')
                arr2.sort(reverse=True)
                print(arr2)
            arrg=[]
            for x in range(1,9):
                arrg.append(sh.cell(row,x).value)
            if valor:
                arrg.append('Q'+arr2[0]+'')
            else:
                arrg.append('N/A')
            linea = '|'.join(arrg)
            tt.write(linea + '\n')



def loaddb():
    BASE_URL = 'http://dx.doi.org/'
    doifile = "PatricioOrio.doi"
    try:
        conn = sqlite3.connect('doidb.db')
        bibfile = doifile.split(".")[0] + ".txt"
        with open(doifile) as dois:
            for line in dois:
                if "doi.org" in line:
                    line = line.split(".org/")[1]
                url = BASE_URL + line
                print(url, end=" JSON ")
                req = urllib.request.Request(url)
                req.add_header('Accept', 'application/json')
                try:
                    print("Connecting!")
                    with urllib.request.urlopen(req, timeout=15) as f:
                        jsonbruto=f.read()
                    sql ='insert into doi(url,json) values(?,?)'
                    cur =conn.cursor()
                    cur.execute(sql,((''.join(url.splitlines())),jsonbruto))
                    conn.commit()
                except Exception as e:
                    print(str(e))
    finally:
        if conn:
            conn.close()

def get_json(url):
    conn = sqlite3.connect('doidb.db')
    cursor=conn.execute('select json from doi where url=?',(url))
    for fila in cursor:
        print(fila)
    return fila

#loaddb()
prepare_file()